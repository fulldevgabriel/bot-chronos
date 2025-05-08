import sqlite3
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
import re

def criar_tabela_horarios():
    with sqlite3.connect('horarios.db') as conn:
        c = conn.cursor()
        c.execute('''CREATE TABLE IF NOT EXISTS horarios (
            id INTEGER PRIMARY KEY,
            usuario TEXT,
            data TEXT,
            hora_entrada TEXT,
            hora_saida TEXT,
            resumo TEXT
        )''')
        conn.commit()

def registrar_entrada(usuario, data, hora):
    with sqlite3.connect('horarios.db') as conn:
        c = conn.cursor()
        c.execute('INSERT INTO horarios (usuario, data, hora_entrada) VALUES (?, ?, ?)', (usuario, data, hora))
        conn.commit()

def registrar_saida(usuario, data, hora, resumo):
    with sqlite3.connect('horarios.db') as conn:
        c = conn.cursor()
        c.execute('UPDATE horarios SET hora_saida=?, resumo=? WHERE usuario=? AND data=?',
                  (hora, resumo, usuario, data))
        conn.commit()

def criar_resumo_excel():
    with sqlite3.connect('horarios.db') as conn:
        c = conn.cursor()
        c.execute("SELECT usuario, data, hora_entrada, hora_saida, resumo FROM horarios")
        dados = c.fetchall()
        colunas = [desc[0] for desc in c.description]
    
    df = pd.DataFrame(dados, columns=colunas)
    wb = Workbook()
    ws = wb.active
    ws.title = "Resumo"
    preencher_worksheet(ws, df)
    formatar_worksheet(ws, df)
    
    nome_arquivo = salvar_arquivo_excel(wb)
    return nome_arquivo

def preencher_worksheet(ws, df):
    for i, col in enumerate(df.columns, 1):
        ws.cell(row=1, column=i, value=col)
    for i, row in enumerate(df.itertuples(), 2):
        ws.cell(row=i, column=1, value=row.usuario)
        data_formatada = formatar_data(row.data)
        ws.cell(row=i, column=2, value=data_formatada)
        ws.cell(row=i, column=3, value=row.hora_entrada[:5] if row.hora_entrada else '')
        ws.cell(row=i, column=4, value=row.hora_saida[:5] if row.hora_saida else 'Pendente')
        ws.cell(row=i, column=5, value=row.resumo)

def formatar_data(data):
    try:
        return datetime.strptime(data, '%Y-%m-%d').strftime('%d/%m')
    except ValueError:
        return data

def formatar_worksheet(ws, df):
    azul_claro = PatternFill(start_color="9FC5E8", end_color="9FC5E8", fill_type="solid")
    verde_claro = PatternFill(start_color="B6D7A8", end_color="B6D7A8", fill_type="solid")
    roxo_claro = PatternFill(start_color="B4A7D6", end_color="B4A7D6", fill_type="solid")
    bold_black_font = Font(bold=True, color="000000")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    aplicar_formatação_colunas(ws, azul_claro, verde_claro, roxo_claro, bold_black_font)
    aplicar_formatação_linhas(ws, azul_claro, verde_claro, roxo_claro)
    ajustar_largura_colunas(ws, df)
    aplicar_bordas(ws, thin_border)

def aplicar_formatação_colunas(ws, azul_claro, verde_claro, roxo_claro, bold_black_font):
    for col in range(1, 6):
        cell = ws.cell(row=1, column=col)
        if col == 1:
            cell.fill = azul_claro
        elif col in [2, 3, 4]:
            cell.fill = verde_claro
        elif col == 5:
            cell.fill = roxo_claro
        cell.font = bold_black_font

def aplicar_formatação_linhas(ws, azul_claro, verde_claro, roxo_claro):
    for row in range(2, ws.max_row + 1):
        ws.cell(row=row, column=1).fill = azul_claro
        ws.cell(row=row, column=2).fill = verde_claro
        ws.cell(row=row, column=3).fill = verde_claro
        ws.cell(row=row, column=4).fill = verde_claro
        ws.cell(row=row, column=5).fill = roxo_claro
        ws.cell(row=row, column=5).alignment = Alignment(wrap_text=True, vertical='top')

def ajustar_largura_colunas(ws, df):
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[col_letter].width = adjusted_width

def aplicar_bordas(ws, thin_border):
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border

def salvar_arquivo_excel(wb):
    nome_arquivo = f'resumo_horário_ponto.xlsx'
    wb.save(nome_arquivo)
    return nome_arquivo
